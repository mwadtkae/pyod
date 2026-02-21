# app.py
import os
import pickle
import sys
import subprocess
from datetime import datetime
from threading import Lock
from typing import Any, Dict, List, Optional, Tuple, Union

from fastapi import FastAPI, HTTPException, Body
from openpyxl import Workbook, load_workbook

# ------------------------------------------------------------
# Unpickle shim:
# Your pickle references "__main__.GroupRule".
# When running under uvicorn, __main__ is uvicorn, not your trainer.
# Define a compatible placeholder so pickle.load can resolve it.
# ------------------------------------------------------------
class GroupRule:
    """Placeholder used ONLY to unpickle legacy models."""
    pass

# Make pickle able to resolve "__main__.GroupRule"
sys.modules["__main__"].GroupRule = GroupRule
# ------------------------------------------------------------

app = FastAPI(title="Timesheet Scoring Service", version="1.0")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DEFAULT_MODEL_PATHS = [
    os.path.join(BASE_DIR, "models", "halfthreshold_model.pkl"),
    os.path.join(BASE_DIR, "halfthreshold_model.pkl"),
]

MODEL_PATH = os.environ.get("MODEL_PATH", "")
MODEL: Optional[dict] = None

EMPLOYEE_KEYS = ["Name", "Employee", "employee", "name"]
PROJECT_KEYS = ["Project Name", "Project_Combined", "project", "project_name", "Project"]
DATE_KEYS = ["Timesheet Date", "TimesheetDate", "date", "Date"]
HOURS_KEYS = ["Time in hours", "Hours", "hours", "time_in_hours"]

# Excel + script runner settings (same folder as app.py by default)
SCORES_XLSX_PATH = os.environ.get("SCORES_XLSX_PATH", os.path.join(BASE_DIR, "scores_log.xlsx"))
RUN_SCRIPT_NAME = os.environ.get("RUN_SCRIPT_NAME", "run_me.py")
RUN_SCRIPT_PATH = os.path.join(BASE_DIR, RUN_SCRIPT_NAME)

_excel_lock = Lock()
_script_lock = Lock()


# -----------------------------
# Helpers (basic parsing)
# -----------------------------
def _first_present(d: Dict[str, Any], keys: List[str]) -> Optional[Any]:
    for k in keys:
        if k in d:
            return d[k]
    return None


def _coerce_float(x: Any) -> float:
    try:
        return float(x)
    except Exception:
        raise ValueError(f"Invalid hours value: {x!r}")


def _normalize_str(x: Any) -> str:
    return str(x).strip()


# -----------------------------
# Rule conversion + scoring
# -----------------------------
def _normalize_rule_obj(rule_obj: Any) -> Dict[str, Any]:
    """
    Convert legacy rule objects (e.g. GroupRule instance) into a plain dict:
      { "threshold": float, "direction": "below|above", "rule": str }
    """
    if isinstance(rule_obj, dict):
        return rule_obj

    if hasattr(rule_obj, "__dict__"):
        d = dict(rule_obj.__dict__)
        out = {
            "threshold": d.get("threshold", d.get("thr", d.get("t"))),
            "direction": d.get("direction", d.get("dir")),
            "rule": d.get("rule", d.get("name", "RULE")),
        }
        if out["threshold"] is None or out["direction"] is None:
            raise ValueError(f"Rule object missing threshold/direction: {d}")
        return out

    raise ValueError(f"Unsupported rule type in pickle: {type(rule_obj)}")


def _rule_flag(rule_dict: Dict[str, Any], hours: float) -> Tuple[bool, str, float, str]:
    direction = str(rule_dict.get("direction", "")).lower().strip()
    if direction not in ("below", "above"):
        raise ValueError(f"Rule direction missing/invalid: {direction!r}")

    try:
        threshold = float(rule_dict["threshold"])
    except KeyError:
        raise ValueError("Rule missing required key: 'threshold'")
    except Exception:
        raise ValueError(f"Rule threshold not numeric: {rule_dict.get('threshold')!r}")

    rule_name = str(rule_dict.get("rule", "RULE"))

    if direction == "below":
        return (hours < threshold), rule_name, threshold, direction
    else:
        return (hours > threshold), rule_name, threshold, direction


def _load_model() -> dict:
    candidates = [MODEL_PATH] if MODEL_PATH else DEFAULT_MODEL_PATHS
    model_file = next((p for p in candidates if p and os.path.exists(p)), None)
    if not model_file:
        raise FileNotFoundError(
            "Could not find model pickle. Set MODEL_PATH env var or place it in one of:\n"
            + "\n".join(DEFAULT_MODEL_PATHS)
        )

    with open(model_file, "rb") as f:
        model = pickle.load(f)

    if not isinstance(model, dict):
        raise ValueError(f"Model pickle must be a dict payload, got: {type(model)}")

    if "rules" not in model or not isinstance(model["rules"], dict):
        raise ValueError("Model payload missing 'rules' dict")

    # Convert legacy rule objects -> dicts
    converted_rules: Dict[Any, Dict[str, Any]] = {}
    for k, v in model["rules"].items():
        converted_rules[k] = _normalize_rule_obj(v)
    model["rules"] = converted_rules

    model["_model_file"] = model_file

    # Optional: write a clean pickle (dict-only rules) so next run won't need GroupRule
    try:
        clean_path = os.path.join(os.path.dirname(model_file), "halfthreshold_model_clean.pkl")
        with open(clean_path, "wb") as wf:
            pickle.dump(model, wf)
        model["_clean_model_file"] = clean_path
    except Exception:
        pass

    return model


def _get_rule_for(emp: str, proj: str) -> Optional[Dict[str, Any]]:
    rules = MODEL.get("rules", {}) if MODEL else {}

    key = (emp, proj)
    if key in rules:
        return rules[key]

    key2 = (_normalize_str(emp), _normalize_str(proj))
    if key2 in rules:
        return rules[key2]

    return None


def _parse_payload(payload: Union[List[Dict[str, Any]], Dict[str, Any]]) -> List[Dict[str, Any]]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        if "entries" in payload and isinstance(payload["entries"], list):
            return payload["entries"]
        return [payload]
    raise HTTPException(status_code=400, detail="Payload must be a JSON object or a list of objects.")


# -----------------------------
# Excel append helpers
# -----------------------------
def _ensure_workbook(path: str):
    """
    Create workbook + header if missing.
    Returns (wb, ws).
    """
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "scores"
    ws.append([
        "timestamp_utc",
        "Name",
        "Project Name",
        "Timesheet Date",
        "Time in hours",
        "flagged",
        "threshold",
        "direction",
        "rule",
    ])
    wb.save(path)
    return wb, ws

from openpyxl import Workbook, load_workbook

REQUIRED_HEADERS = ["Name", "Project Name", "Timesheet Date", "Time in hours"]

def _ensure_workbook(path: str):
    """
    Ensures workbook exists AND required headers exist somewhere on row 1.
    Does NOT reorder existing columns; it appends missing headers to the end.
    Returns (wb, ws).
    """
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "scores"
        ws.append([])  # create row 1

    # Build existing header map from row 1
    header_map: Dict[str, int] = {}
    max_col = ws.max_column if ws.max_column else 0

    for col in range(1, max_col + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        header_map[str(v).strip()] = col

    # Add any missing required headers at end (don’t move existing)
    for h in REQUIRED_HEADERS:
        if h not in header_map:
            max_col += 1
            ws.cell(row=1, column=max_col, value=h)
            header_map[h] = max_col

    wb.save(path)
    return wb, ws

from typing import Any, Dict, List

def _append_scores_to_excel(rows: List[Dict[str, Any]], path: str) -> int:
    """
    Appends ONLY 4 inputs into their respectively named columns:
      Name -> "Name"
      Project Name -> "Project Name"
      Timesheet Date -> "Date"
      Time in hours -> "Time in hours"

    Writes by header lookup (row 1), so it won't shift even if sheet has extra columns.
    """
    with _excel_lock:
        wb, ws = _ensure_workbook(path)

        # Rebuild header map (after ensure)
        header_map: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=col).value
            if v is None:
                continue
            header_map[str(v).strip()] = col

        appended = 0
        next_row = ws.max_row + 1 if ws.max_row else 2

        for r in rows:
            name = r.get("Name")
            project = r.get("Project Name")
            date_val = r.get("Timesheet Date")
            hours = r.get("Time in hours", r.get("hours"))

            # Require all 4 fields (so you don't write broken rows)
            if name is None or project is None or date_val is None or hours is None:
                continue

            ws.cell(row=next_row, column=header_map["Name"], value=str(name))
            ws.cell(row=next_row, column=header_map["Project Name"], value=str(project))
            ws.cell(row=next_row, column=header_map["Timesheet Date"], value=str(date_val))
            ws.cell(row=next_row, column=header_map["Time in hours"], value=float(hours))

            appended += 1
            next_row += 1

        wb.save(path)
        return appended



# -----------------------------
# Startup
# -----------------------------
@app.on_event("startup")
def startup():
    global MODEL
    try:
        MODEL = _load_model()
        print(f"Loaded model: {MODEL.get('_model_file')}")
        if MODEL.get("_clean_model_file"):
            print(f"Wrote clean model: {MODEL.get('_clean_model_file')}")
        print(f"Rules loaded: {len(MODEL.get('rules', {}))}")
    except Exception as e:
        raise RuntimeError(f"Failed to load model pickle: {e}") from e


# -----------------------------
# Endpoints
# -----------------------------
@app.get("/health")
def health():
    if not MODEL:
        return {"status": "error", "detail": "model not loaded"}
    return {
        "status": "ok",
        "model_file": MODEL.get("_model_file"),
        "clean_model_file": MODEL.get("_clean_model_file"),
        "type": MODEL.get("type", "unknown"),
        "rules_count": len(MODEL.get("rules", {})),
        "scores_excel": SCORES_XLSX_PATH,
        "run_script": RUN_SCRIPT_PATH,
    }


@app.post("/score_entries")
def score_entries(payload: Union[List[Dict[str, Any]], Dict[str, Any]] = Body(...)):
    if not MODEL:
        raise HTTPException(status_code=500, detail="Model not loaded")

    entries = _parse_payload(payload)
    results: List[Dict[str, Any]] = []

    for i, row in enumerate(entries):
        emp = _first_present(row, EMPLOYEE_KEYS)
        proj = _first_present(row, PROJECT_KEYS)
        dt = _first_present(row, DATE_KEYS)
        hrs = _first_present(row, HOURS_KEYS)

        if emp is None or proj is None or hrs is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Entry index {i} missing required fields. "
                    f"Need employee/name, project, hours. Got keys: {list(row.keys())}"
                ),
            )

        emp_s = _normalize_str(emp)
        proj_s = _normalize_str(proj)

        try:
            hours_f = _coerce_float(hrs)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"Entry index {i}: {e}")

        rule = _get_rule_for(emp_s, proj_s)
        if rule is None:
            results.append(
                {
                    "index": i,
                    "Name": emp_s,
                    "Project Name": proj_s,
                    "Timesheet Date": dt,
                    "Time in hours": hours_f,
                    "flagged": False,
                    "status": "NO_RULE",
                }
            )
            continue

        try:
            flagged, rule_name, threshold, direction = _rule_flag(rule, hours_f)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Entry index {i}: scoring failed: {e}")

        # Minimal score output, plus threshold info so you can demo cut points
        results.append(
            {
                "index": i,
                "Name": emp_s,
                "Project Name": proj_s,
                "Timesheet Date": dt,
                "Time in hours": hours_f,
                "flagged": flagged,
                "rule": rule_name,
            }
        )

    return {"count": len(results), "results": results}


@app.post("/append_scores")
def append_scores(payload: Union[List[Dict[str, Any]], Dict[str, Any]] = Body(...)):
    """
    Accepts:
      - a JSON list of rows
      - OR an object with {"results": [...]}
    This is designed so you can pipe the response of /score_entries directly into it.
    """
    if isinstance(payload, dict) and "results" in payload and isinstance(payload["results"], list):
        rows = payload["results"]
    elif isinstance(payload, list):
        rows = payload
    else:
        raise HTTPException(status_code=400, detail="Send a JSON list of rows or an object with {results: [...]}")

    try:
        appended = _append_scores_to_excel(rows, SCORES_XLSX_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to append to Excel: {e}")

    return {"status": "ok", "appended": appended, "excel_path": SCORES_XLSX_PATH}


@app.post("/run_script")
def run_script():
    """
    Runs a local python script in the same folder as app.py.
    Default script name: run_me.py (override with RUN_SCRIPT_NAME env var)
    """
    if not os.path.exists(RUN_SCRIPT_PATH):
        raise HTTPException(
            status_code=404,
            detail=f"Script not found at {RUN_SCRIPT_PATH}. Create it or set RUN_SCRIPT_NAME env var.",
        )

    if not _script_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Script is already running")

    try:
        proc = subprocess.run(
            [sys.executable, RUN_SCRIPT_PATH],
            cwd=BASE_DIR,
            capture_output=True,
            text=True,
        )
        return {
            "status": "ok" if proc.returncode == 0 else "error",
            "returncode": proc.returncode,
            "stdout": proc.stdout,
            "stderr": proc.stderr,
            "script": RUN_SCRIPT_NAME,
        }
    finally:
        _script_lock.release()
